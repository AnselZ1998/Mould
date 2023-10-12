# -*-coding: Utf-8 -*-
# @File : RollingExcel_Test.py
# author : Ansel Z
# Time : 2023/10/11 20:47
# Version :1.0.0
#
#设置单元格内的字符
    #判断中英文，中文标红，英文标黄
    #中文字体设置，英文字体设置
    #中文下对齐，英文上对齐

import openpyxl
import re
from tqdm import tqdm
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
Image.MAX_IMAGE_PIXELS = None


class Adjust_Excel():

    def __init__(self):
        self.work_excel = None
        self.cell_v = None
        self.cont = None
        self.font_EN = None
        self.font_CH = None
        self.red_fill = None
        self.yellow_fill = None
        self.bottom_align = None
        self.top_align = None

    def LoadExcel(self):
        excel_path = '/Users/anselz/PROJECT/Python_Project_Z/Ansel/PROJECT/RollingExcel/Test/0921片尾字幕Final.xlsx'
        self.work_excel = openpyxl.load_workbook(excel_path)
        #选中sheet表格
        sht = self.work_excel["Sheet3"]
        #选中单元格区域
        self.cells = sht["A1:BI1015"]

        for row in sht.iter_rows():
            sht.row_dimensions[row[0].row].height = 20
        for column in sht.iter_cols():
            sht.column_dimensions[column[0].column_letter].width = 4

    def WorldFont(self):
        #定义字体样式

        self.font_CN = Font(name="微软雅黑", size=16, bold=False, italic=False, color="FF0000")
        self.font_EN = Font(name="微软雅黑", size=10, bold=False, italic=False, color="FF0000")
        self.red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.top_align = Alignment(vertical='top')
        self.bottom_align = Alignment(vertical='bottom')

    def CNorEN(self):

            return bool(re.search(r'[\u4e00-\u9fa5]', self.cont))

    def Forexel(self):

        for i in tqdm(self.cells):
            for self.cell_v in i:
                if self.cell_v.value is not None:
                    self.cont = str(self.cell_v.value)
                    #print(self.cont)
                    if self.CNorEN():
                        #print('CN')
                        self.cell_v.font = self.font_CN
                        self.cell_v.fill = self.red_fill
                        self.cell_v.alignment = self.bottom_align
                    else:
                        self.cell_v.font = self.font_EN
                        self.cell_v.fill = self.yellow_fill
                        self.cell_v.alignment = self.top_align
                        #print('EN')


              #  cell.font = font
              #  print(cell.value)

        self.work_excel.save(filename="xiugaihou.xlsx")


if __name__ == '__main__':
    adjust_excel = Adjust_Excel()
    #print(adjust_excel.CNorEN())
    adjust_excel.LoadExcel()
    adjust_excel.WorldFont()

    adjust_excel.Forexel()